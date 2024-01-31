import mysql.connector 
from datetime import datetime


class Database:
    #1
    def __init__(self):
        self.db = self.ulanish()
        self.cursor = self.db.cursor()

    def ulanish(self):
        return mysql.connector.connect(
            host="localhost",
            user="root",
            password="root",
            database="oquv_markazi"
        )
    
    def ishlatish(self, sql, fetchall=False, fetchone=False, commit=False):
        db = self.ulanish()
        cursor = self.db.cursor()
        data = None
        if fetchall:
            cursor.execute(sql)
            data = cursor.fetchall()
        elif fetchone:
            cursor.execute(sql)
            data = cursor.fetchone()
        elif commit:
            cursor.execute(sql)
            db.commit()

        db.close()
        return data

    #2
    def kurslar_nomi(self):
        sql = ("SELECT kurs_nomi FROM kurslar")
        return self.ishlatish(sql,fetchall=True)
    
    #3
    def kurs_nomi(self,id):
        sql = (f"SELECT * FROM kurslar WHERE id = {id}")
        return self.ishlatish(sql,fetchone=True)
    
    #4
    def uqituvchilar_fio(self):
        sql = ("SELECT ism,familiya FROM uqituvchilar")
        return self.ishlatish(sql,fetchall=True)
    
    #5
    def search_uquvchi(self,suz):
        sql = (f'SELECT ism FROM uquvchilar WHERE ism LIKE "%{suz}"')
        return self.ishlatish(sql,fetchall=True)
    
    #6
    def count_kurs_uquvchilar(self,kurs_nomi):
        try:
            sql = (f"SELECT * FROM uquvchilar WHERE id IN(SELECT uquvchi_id FROM uquvchi_kurslari WHERE kurs_id IN (SELECT id FROM kurslar WHERE kurs_nomi = '{kurs_nomi}'))")
            return self.ishlatish(sql,fetchall=True)
        except Exception:
            return "Kurs_uquvchilar metodida xatolik !!!"

    #7    
    def uqituvchi_kurslari(self,ism,familiya):
        sql = (f"SELECT * FROM kurslar WHERE id IN(SELECT kurs_id FROM uqituvchi_kurslari WHERE uqituvchi_id IN (SELECT id FROM uqituvchilar WHERE ism='{ism}' and familiya='{familiya}'))")
        return self.ishlatish(sql,fetchall=True)

    #8    
    def uqituvchi_uquvchilari(self,ism,familiya):
        sql = (f"SELECT ism,familiya FROM uquvchilar WHERE id IN(SELECT uquvchi_id FROM uquvchi_kurslari WHERE uqituvchi_id IN (SELECT id FROM uqituvchilar WHERE ism ='{ism}' AND familiya='{familiya}'))")
        return self.ishlatish(sql,fetchall=True)
    
    #9
    def uquvchi_uqituvchisi(self,ism,familiya):
        sql = (f"SELECT ism,familiya FROM uqituvchilar WHERE id IN (SELECT uqituvchi_id FROM uquvchi_kurslari WHERE uquvchi_id IN (SELECT id FROM uquvchilar WHERE ism='{ism}' and familiya='{familiya}'))") 
        return self.ishlatish(sql,fetchall=True)

    #10
    def kurs_uquvchilari(self,id):
        sql = (f"SELECT COUNT(uquvchi_id) FROM uquvchi_kurslari WHERE kurs_id = {id}")
        return self.ishlatish(sql,fetchone=True)
    
    #11
    def count_tulovlar(self,sana):
        sql=f"SELECT COUNT(tulov) FROM tulovlar WHERE vaqti >='{sana} 00:00:00' AND vaqti<='{sana} 23:59:59'"
        return self.ishlatish(sql,fetchone=True)

    #12
    def sum_tulovlar(self,sana):
        sql=f"SELECT SUM(tulov) FROM tulovlar WHERE vaqti >='{sana} 00:00:00' AND vaqti<='{sana} 23:59:59'"
        return self.ishlatish(sql,fetchone=True)

    #13
    def sum_sanalar_tulovi(self,sana1,sana2):
        sql = f"SELECT SUM(tulov) FROM tulovlar WHERE vaqti BETWEEN '{sana1} 00:00:00' AND '{sana2} 23:59:59'"
        return self.ishlatish(sql, fetchone=True)

    #14
    def sum_uqituvchi_tulovlari(self,id,sana1,sana2):
        sql=f"SELECT SUM(tulov) FROM tulovlar WHERE vaqti BETWEEN '{sana1} 00:00:00' AND '{sana2} 23:59:59' AND uquvchi_id IN(SELECT uquvchi_id FROM uquvchi_kurslari WHERE uqituvchi_id={id})"
        return self.ishlatish(sql,fetchone=True)

    #15
    def sum_uquvchi_tulovlari(self,id):
        sql=f"SELECT SUM(tulov) FROM tulovlar WHERE uquvchi_id={id}"
        return self.ishlatish(sql,fetchone=True)

    #16
    def kurs_narxini_oshirish(self,id1,id2,foiz):
        if foiz<0:
            sql = (f"UPDATE kurslar SET narxi = narxi - narxi*{foiz}/100' WHERE id BETWEEN {id1} AND {id2}")
        if foiz>0:
            sql = (f"UPDATE kurslar SET narxi = narxi + narxi*{foiz}/100 WHERE id BETWEEN {id1} AND {id2}")
        self.ishlatish(sql,commit=True)

    #17
    def insert_kurslar(self,kurs_nomi,narxi,davomiyligi):
        sql = f"INSERT INTO kurslar(kurs_nomi,narxi,davomiyligi) VALUES('{kurs_nomi}',{narxi},'{davomiyligi}')"
        self.ishlatish(sql,commit=True)

    #18
    def insert_uquvchilar(self,ism,familiya,t_sana,tel):
        sql = (f"INSERT INTO uquvchilar(ism,familiya,tugilgan_sana,tel_raqam) VALUES('{ism}','{familiya}','{t_sana}','{tel}')")
        self.ishlatish(sql,commit=True)

    #19
    def insert_uqituvchilar(self,ism,familiya,yunalishi,tel):
        sql = (f"INSERT INTO uqituvchilar(ism,familiya,yunalishi,tel_raqami) VALUES('{ism}','{familiya}','{yunalishi}','{tel}')")
        self.ishlatish(sql,commit=True)

    def uquvchi_id(self,ism, familiya):
        sql = f"""
                SELECT id 
                FROM uquvchilar
                WHERE ism = '{ism}' AND familiya = '{familiya}'
                """
        return self.ishlatish(sql, fetchone=True)

    def kurs_id(self, kurs_nomi):
        sql = f"""
                SELECT id 
                FROM kurslar
                WHERE kurs_nomi = '{kurs_nomi}'
        """
        return self.ishlatish(sql, fetchone=True)

    #20
    def insert_tulov(self, ism, familiya, tulov, kurs_nomi):
        UquvchiId = self.uquvchi_id(ism, familiya)
        KursId = self.kurs_id(kurs_nomi)
        sana = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sql = f"""
            INSERT INTO tulovlar 
            VALUES
            ({UquvchiId[0]}, {tulov}, {KursId[0]}, '{sana}')
        """
        self.ishlatish(sql, commit=True)
        
    #21
    def change_dars_vaqti(self,ism,familiya,new_vaqt):
        sql = (f"UPDATE uqituvchi_kurslari SET vaqti = '{new_vaqt}' WHERE uqituvchi_id IN(SELECT id FROM uqituvchilar WHERE ism = '{ism}' AND familiya='{familiya}') ")
        self.ishlatish(sql,commit=True)
        
    #22
    def SqlToJadval(self, nomi):
        sql = f"SELECT * FROM {nomi}"
        data = self.ishlatish(sql, fetchall=True)
        excel_table = Excel()
        excel_table.write_table(data)

    #23
    def TableToSql(self,fayl_nomi=None):
        excel = Excel(f"{fayl_nomi}.xlsx","Sheet")
        a = excel.read_text()
        # b = ""
        # for i in a:
        #     b+=str(i)+","
        # print(b)
        a = excel.read_text()
        b = ",".join(map(str, a))
        print(b)
        sql = f"INSERT INTO yangi_uquvchilar(id,ism,familiya,t_sana,t_raqam) VALUES({b})"
        self.ishlatish(sql,commit=True)

from openpyxl import Workbook, load_workbook

class Excel:
    def __init__(self, name=None, list_nomi=None):
    # def __init__(self, name=None):
        if name:
            self.file = load_workbook(name)
            self.jadval = self.file[list_nomi]
        else:
            self.file = Workbook()
            self.jadval = self.file.active

    def read_text(self):
        l = []
        for i in self.jadval.iter_rows(min_row=1,max_row=self.jadval.max_row, values_only=True):  
            l.append(i)
        return l

    def write_table(self, data):
        # self.jadval.append(["№", "ism", "familiya", "T_sana", "Tel"])
        for i in data:
            self.jadval.append(i)
        self.saqlash("uquvchilar.xlsx")

    def saqlash(self, manzil=None):
        self.file.save(manzil)