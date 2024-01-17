from openpyxl import Workbook

class Excel:
    def __init__(self,name):
        self.name = name
        self.excel = Workbook()
        self.file = self.excel.active

    def read_text(self):
        for i in self.file.iter_rows(min_row=1,max_row=self.file.max_row, values_only=True):  
            print(i)

    def write_cell(self,adress,qiymat):
        try:
            self.file[f"{adress}"].value = qiymat
        except Exception as q:
            print(f"{q}")

    def add(self,*qiymat):
        try:
            self.file.append(qiymat)
        except Exception as q:
            print(f"{q}")

    def qushish(self,adress1,adress2,adress3):
        try:
           a = self.file[adress1].value
           b = self.file[adress2].value
           self.file[adress3] = f"{a+" "}{b}"
        except Exception as q:
            print(f"{q}")

    def massiv(self,adress1,adress2):
        a = int(adress1[1:])
        b = int(adress2[1:])
        c = []
        for i in self.file.iter_rows(min_row=a,max_row=b,values_only=True):
            c.append(i)
        print(c)

    def del_cell(self,adress):
        self.file[adress].value = None 

    def __repr__(self):
        return self.name
    
    def __call__(self,adress):
        return self.file[adress].value

    def saqlash(self,manzil):
        self.excel.save(self.name)

